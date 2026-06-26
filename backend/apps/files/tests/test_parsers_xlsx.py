from io import BytesIO

import pytest
from openpyxl import Workbook

from apps.files.parsers import UnsupportedFormatError, parse_upload, supported_upload_accept
from apps.files.parsers.xlsx import XlsxParser


def build_sample_xlsx_bytes(rows=None) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ventas"
    data = rows or [
        ["Region", "Revenue"],
        ["EMEA", 100],
        ["APAC", 200],
    ]
    for row in data:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


@pytest.mark.django_db
class TestXlsxParser:
    def test_parse_simple_sheet(self):
        parsed = parse_upload(build_sample_xlsx_bytes(), "ventas.xlsx")
        assert parsed.format_key == "xlsx"
        assert parsed.content_json["format"] == "xlsx"
        assert parsed.content_json["source"] == "upload"
        assert parsed.content_json["title"] == "ventas"
        assert len(parsed.content_json["sheets"]) == 1
        sheet = parsed.content_json["sheets"][0]
        assert sheet["headers"] == ["Region", "Revenue"]
        assert len(sheet["rows"]) == 2
        assert parsed.preview_html

    def test_parse_multiple_sheets(self):
        workbook = Workbook()
        first = workbook.active
        first.title = "One"
        first.append(["A", "B"])
        first.append(["1", "2"])
        second = workbook.create_sheet("Two")
        second.append(["X", "Y"])
        second.append(["9", "8"])
        buffer = BytesIO()
        workbook.save(buffer)
        parsed = parse_upload(buffer.getvalue(), "multi.xlsx")
        assert len(parsed.content_json["sheets"]) == 2

    def test_parse_truncates_rows(self):
        rows = [["Col1", "Col2"]]
        rows.extend([[str(i), str(i + 1)] for i in range(60)])
        parsed = parse_upload(build_sample_xlsx_bytes(rows), "big.xlsx")
        sheet = parsed.content_json["sheets"][0]
        assert len(sheet["rows"]) == 50
        assert parsed.content_json.get("parse_warnings")

    def test_invalid_file_raises(self):
        with pytest.raises(ValueError, match="Invalid Excel file"):
            parse_upload(b"not-an-xlsx", "bad.xlsx")

    def test_unsupported_format_raises(self):
        with pytest.raises(UnsupportedFormatError):
            parse_upload(b"hello", "notes.txt", "text/plain")

    def test_supports_by_extension(self):
        parser = XlsxParser()
        assert parser.supports("", "report.xlsx") is True
        assert parser.supports("", "report.csv") is False

    def test_supported_upload_accept_includes_xlsx(self):
        accept = supported_upload_accept()
        assert ".xlsx" in accept
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in accept
