from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from apps.agent.tools.spreadsheet import build_preview_html
from apps.agent.tools.spreadsheet_content import (
    MAX_COLUMNS,
    MAX_SHEETS,
    MAX_TABLE_ROWS,
    validate_xlsx_content_json,
)
from apps.files.models import XLSX_MIME
from apps.files.parsers.base import ParsedUpload

XLSX_EXTENSIONS = frozenset({".xlsx"})


def _cell_to_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _row_is_empty(values: list[str]) -> bool:
    return not any(values)


def _title_from_filename(original_name: str) -> str:
    stem = Path(original_name).stem.strip()
    if stem:
        return stem
    return "Hoja importada"


def _parse_sheet(worksheet, warnings: list[str]) -> dict | None:
    rows_raw: list[list[str]] = []
    for row in worksheet.iter_rows(values_only=True):
        values = [_cell_to_str(cell) for cell in row]
        if not _row_is_empty(values):
            rows_raw.append(values)

    if not rows_raw:
        return None

    headers = rows_raw[0]
    if len(headers) > MAX_COLUMNS:
        headers = headers[:MAX_COLUMNS]
        warnings.append(
            f"Hoja '{worksheet.title}': columnas truncadas a {MAX_COLUMNS}"
        )

    data_rows = rows_raw[1:]
    if len(data_rows) > MAX_TABLE_ROWS:
        data_rows = data_rows[:MAX_TABLE_ROWS]
        warnings.append(
            f"Hoja '{worksheet.title}': filas truncadas a {MAX_TABLE_ROWS}"
        )

    normalized_rows = []
    for row in data_rows:
        padded = list(row[: len(headers)])
        if len(padded) < len(headers):
            padded.extend([""] * (len(headers) - len(padded)))
        normalized_rows.append(padded)

    return {
        "name": worksheet.title,
        "headers": headers,
        "rows": normalized_rows,
    }


class XlsxParser:
    extensions = XLSX_EXTENSIONS
    mime_types = frozenset({XLSX_MIME})

    def supports(self, mime_type: str, original_name: str) -> bool:
        if mime_type == XLSX_MIME:
            return True
        return Path(original_name).suffix.lower() in XLSX_EXTENSIONS

    def parse(self, file_bytes: bytes, original_name: str) -> ParsedUpload:
        warnings: list[str] = []
        try:
            workbook = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
        except Exception as exc:
            raise ValueError(f"Invalid Excel file: {exc}") from exc

        sheets: list[dict] = []
        try:
            for worksheet in workbook.worksheets:
                if worksheet.sheet_state != "visible":
                    continue
                parsed = _parse_sheet(worksheet, warnings)
                if parsed is not None:
                    sheets.append(parsed)
                if len(sheets) >= MAX_SHEETS:
                    if len(workbook.worksheets) > MAX_SHEETS:
                        warnings.append(f"Hojas truncadas a {MAX_SHEETS}")
                    break
        finally:
            workbook.close()

        if not sheets:
            raise ValueError("Excel file contains no readable data")

        title = _title_from_filename(original_name)
        content_json = validate_xlsx_content_json(title, sheets)
        content_json["source"] = "upload"
        content_json["role"] = "context"
        if warnings:
            content_json["parse_warnings"] = warnings

        preview_html = build_preview_html(content_json)

        return ParsedUpload(
            content_json=content_json,
            preview_html=preview_html,
            mime_type=XLSX_MIME,
            format_key="xlsx",
            parse_warnings=warnings,
        )
