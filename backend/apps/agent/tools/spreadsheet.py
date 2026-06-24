import json
import re
from io import BytesIO
from typing import Annotated

from langchain_core.tools import InjectedToolCallId, tool
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from apps.agent.cancellation import check_agent_not_cancelled
from apps.agent.context import get_agent_conversation, get_agent_user
from apps.agent.tools.errors import build_tool_error_response
from apps.agent.tools.spreadsheet_content import (
    revalidate_xlsx_content_json,
    validate_xlsx_content_json,
)
from apps.files.models import XLSX_MIME
from apps.files.services import escape_preview_text as esc
from apps.files.services import (
    get_file_for_conversation,
    save_generated_file,
    serialize_file_for_agent,
    serialize_file_for_ui,
    update_generated_file,
)

_SPREADSHEET_DISPLAY_REGISTRY: dict[str, dict] = {}

AGENT_INSTRUCTION_AFTER_SPREADSHEET = (
    "La hoja de cálculo ya está visible en el chat del usuario. "
    "NO repitas el contenido de la tabla en tu siguiente mensaje. "
    "Solo añade una frase breve si aporta contexto, o termina sin texto."
)

TONE_FONT_COLORS = {
    "success": "166534",
    "danger": "B91C1C",
    "warning": "B45309",
    "muted": "6B7280",
}

ALIGN_MAP = {
    "left": "left",
    "right": "right",
    "center": "center",
}


def pop_spreadsheet_display(tool_call_id: str | None) -> dict | None:
    if not tool_call_id:
        return None
    return _SPREADSHEET_DISPLAY_REGISTRY.pop(tool_call_id, None)


def _sanitize_filename(name: str, fallback: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*]', "", name).strip()
    if not cleaned:
        cleaned = fallback
    if not cleaned.lower().endswith(".xlsx"):
        cleaned = f"{cleaned}.xlsx"
    return cleaned[:200]


def validate_content_json(title: str, sheets: list) -> dict:
    return validate_xlsx_content_json(title, sheets)


def _col_letter(index: int) -> str:
    letters = ""
    n = index + 1
    while n > 0:
        n, rem = divmod(n - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def _apply_cell_style(cell, cell_data: dict, *, header: bool = False) -> None:
    bold = header or cell_data.get("bold", False)
    align = cell_data.get("align", "left")
    tone = cell_data.get("tone", "default")
    font_kwargs = {"bold": bold}
    if tone in TONE_FONT_COLORS:
        font_kwargs["color"] = TONE_FONT_COLORS[tone]
    cell.font = Font(**font_kwargs)
    cell.alignment = Alignment(horizontal=ALIGN_MAP.get(align, "left"))


def _write_sheet(ws, sheet_data: dict) -> None:
    headers = sheet_data.get("headers", [])
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        _apply_cell_style(cell, {"align": "left", "bold": True}, header=True)
    for row_idx, row in enumerate(sheet_data.get("rows", []), start=2):
        cells = row.get("cells", [])
        row_style = row.get("style", "default")
        for col_idx, cell_data in enumerate(cells[: len(headers)], start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_data.get("value", ""))
            styled = dict(cell_data)
            if row_style in ("total", "subtotal") and not styled.get("bold"):
                styled["bold"] = True
            _apply_cell_style(cell, styled)


def build_xlsx(content_json: dict) -> bytes:
    content_json = revalidate_xlsx_content_json(content_json)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_data in content_json.get("sheets", []):
        ws = workbook.create_sheet(title=sheet_data["name"][:31])
        _write_sheet(ws, sheet_data)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _cell_classes(cell_data: dict, *, header: bool = False, row_style: str = "default") -> str:
    classes = ["ay-sheet-preview__cell"]
    if header:
        classes.append("ay-sheet-preview__cell--header")
    else:
        align = cell_data.get("align", "left")
        if align != "left":
            classes.append(f"ay-sheet-preview__cell--{align}")
        tone = cell_data.get("tone", "default")
        if tone != "default":
            classes.append(f"ay-sheet-preview__cell--{tone}")
        if cell_data.get("bold") or row_style in ("total", "subtotal"):
            classes.append("ay-sheet-preview__cell--bold")
        if row_style in ("total", "subtotal"):
            classes.append(f"ay-sheet-preview__cell--{row_style}")
    return " ".join(classes)


def _grid_template_cols(col_count: int) -> str:
    if col_count <= 0:
        return "38px"
    return f"38px repeat({col_count}, minmax(72px, max-content))"


def _render_sheet_html(sheet_data: dict, *, active: bool) -> str:
    headers = sheet_data.get("headers", [])
    rows = sheet_data.get("rows", [])
    col_count = len(headers)
    col_letters = [_col_letter(i) for i in range(col_count)]
    hidden = "" if active else ' hidden aria-hidden="true"'
    parts = [
        f'<div class="ay-sheet-preview__pane"{hidden} data-sheet-name="{esc(sheet_data["name"])}">',
        '<div class="ay-sheet-preview__formula-bar">',
        '<div class="ay-sheet-preview__formula-ref">A1</div>',
        '<div class="ay-sheet-preview__formula-fx">fx</div>',
        '<div class="ay-sheet-preview__formula-value"></div>',
        "</div>",
        '<div class="ay-sheet-preview__scroll">',
        '<div class="ay-sheet-preview__canvas" aria-hidden="true"></div>',
        (
            f'<div class="ay-sheet-preview__grid" data-col-count="{col_count}" '
            f'data-data-rows="{len(rows)}" '
            f'data-cols-template="{esc(_grid_template_cols(col_count))}" '
            f'style="--ay-sheet-cols: {_grid_template_cols(col_count)}">'
        ),
        '<div class="ay-sheet-preview__corner"></div>',
    ]
    for letter in col_letters:
        parts.append(f'<div class="ay-sheet-preview__col-hdr">{esc(letter)}</div>')
    parts.append('<div class="ay-sheet-preview__row-hdr">1</div>')
    for header in headers:
        header_classes = _cell_classes({"align": "left", "bold": True}, header=True)
        parts.append(
            f'<div class="{header_classes}">'
            f"{esc(header)}</div>"
        )
    for row_idx, row in enumerate(rows, start=2):
        row_style = row.get("style", "default")
        parts.append(f'<div class="ay-sheet-preview__row-hdr">{row_idx}</div>')
        cells = row.get("cells", [])
        for col_idx in range(col_count):
            cell_data = cells[col_idx] if col_idx < len(cells) else {"value": ""}
            parts.append(
                f'<div class="{_cell_classes(cell_data, row_style=row_style)}">'
                f'{esc(cell_data.get("value", ""))}</div>'
            )
    parts.extend(["</div>", "</div>", "</div>"])
    return "".join(parts)


def build_preview_html(content_json: dict) -> str:
    content_json = revalidate_xlsx_content_json(content_json)
    sheets = content_json.get("sheets", [])
    parts = ['<div class="ay-sheet-preview">']
    for idx, sheet_data in enumerate(sheets):
        parts.append(_render_sheet_html(sheet_data, active=idx == 0))
    parts.append('<div class="ay-sheet-preview__tabs">')
    for idx, sheet_data in enumerate(sheets):
        active_class = " ay-sheet-preview__tab--active" if idx == 0 else ""
        parts.append(
            f'<button type="button" class="ay-sheet-preview__tab{active_class}" '
            f'data-sheet-tab="{esc(sheet_data["name"])}">{esc(sheet_data["name"])}</button>'
        )
    parts.extend(["</div>", "</div>"])
    return "".join(parts)


def preview_html_for_file(content_json: dict | None, preview_html: str) -> str:
    if content_json and content_json.get("format") == "xlsx":
        if content_json.get("sheets") is not None:
            return build_preview_html(content_json)
    return preview_html or ""


def _merge_content_json(existing: dict, title, sheets) -> dict:
    merged = dict(existing)
    if title is not None:
        merged["title"] = str(title).strip()
    if sheets is not None:
        merged["sheets"] = sheets
    return validate_content_json(
        merged.get("title", ""),
        merged.get("sheets", []),
    )


def _build_agent_tool_response(file_obj, action: str) -> str:
    return json.dumps(
        {
            "ok": True,
            "action": action,
            "file_id": str(file_obj.id),
            "name": file_obj.original_name,
            "version": file_obj.version,
            "agent_instruction": AGENT_INSTRUCTION_AFTER_SPREADSHEET,
        }
    )


def _register_display(tool_call_id: str, file_obj, updated: bool = False) -> None:
    payload = serialize_file_for_ui(file_obj)
    payload["updated"] = updated
    _SPREADSHEET_DISPLAY_REGISTRY[tool_call_id] = payload


@tool
def create_spreadsheet(
    title: str,
    sheets: list[dict],
    filename: str = "",
    tool_call_id: Annotated[str, InjectedToolCallId] = "",
) -> str:
    """Create an Excel spreadsheet (.xlsx) for the user in the chat.

    Use when the user asks for a spreadsheet, Excel export, or tabular data file.
    Each sheet has a `name`, `headers`, and `rows`. Rows can be plain lists or objects
    with `cells` and optional `style` (total, subtotal). Cells support `value`, `align`,
    `tone` (success, danger, warning, muted), and `bold`.
    The spreadsheet appears in the chat; do not repeat its content in your text response.

    To modify an existing spreadsheet later, use update_spreadsheet with the same file_id.
    """
    check_agent_not_cancelled()
    conversation = get_agent_conversation()
    user = get_agent_user()
    if conversation is None or user is None:
        return build_tool_error_response("No conversation context")

    try:
        content_json = validate_content_json(title, sheets)
    except ValueError as exc:
        return build_tool_error_response(str(exc))

    try:
        original_name = _sanitize_filename(filename, content_json["title"])
        xlsx_bytes = build_xlsx(content_json)
        preview = build_preview_html(content_json)
        file_obj = save_generated_file(
            conversation=conversation,
            user=user,
            original_name=original_name,
            content_json=content_json,
            file_bytes=xlsx_bytes,
            preview_html=preview,
            mime_type=XLSX_MIME,
        )
    except Exception as exc:
        return build_tool_error_response(str(exc))

    _register_display(tool_call_id, file_obj, updated=False)
    return _build_agent_tool_response(file_obj, "created")


@tool
def get_spreadsheet(file_id: str) -> str:
    """Read the full structured content of a spreadsheet by file_id.

    Call this before update_spreadsheet when you need the current sheets.
    """
    conversation = get_agent_conversation()
    if conversation is None:
        return build_tool_error_response("No conversation context")

    file_obj = get_file_for_conversation(file_id, conversation)
    if file_obj is None:
        return build_tool_error_response("Spreadsheet not found in this conversation")
    if file_obj.format_key != "xlsx":
        return build_tool_error_response("File is not a spreadsheet")

    return json.dumps(
        {
            "ok": True,
            **serialize_file_for_agent(file_obj),
            "content_json": file_obj.content_json,
        }
    )


@tool
def update_spreadsheet(
    file_id: str,
    title: str | None = None,
    sheets: list[dict] | None = None,
    tool_call_id: Annotated[str, InjectedToolCallId] = "",
) -> str:
    """Update an existing spreadsheet by file_id.

    Provide only the fields you want to change. sheets replaces all sheets when provided.
    """
    check_agent_not_cancelled()
    conversation = get_agent_conversation()
    if conversation is None:
        return build_tool_error_response("No conversation context")

    file_obj = get_file_for_conversation(file_id, conversation)
    if file_obj is None:
        return build_tool_error_response("Spreadsheet not found in this conversation")
    if file_obj.format_key != "xlsx":
        return build_tool_error_response("File is not a spreadsheet")

    try:
        content_json = _merge_content_json(file_obj.content_json, title, sheets)
    except ValueError as exc:
        return build_tool_error_response(str(exc))

    try:
        xlsx_bytes = build_xlsx(content_json)
        preview = build_preview_html(content_json)
        file_obj = update_generated_file(
            file_obj=file_obj,
            content_json=content_json,
            file_bytes=xlsx_bytes,
            preview_html=preview,
        )
    except Exception as exc:
        return build_tool_error_response(str(exc))

    _register_display(tool_call_id, file_obj, updated=True)
    return _build_agent_tool_response(file_obj, "updated")
