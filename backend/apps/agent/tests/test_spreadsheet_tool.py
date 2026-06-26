import json
from io import BytesIO
from zipfile import ZipFile

import pytest
from django.contrib.auth import get_user_model
from openpyxl import load_workbook

from apps.agent.context import set_agent_context
from apps.agent.tools.spreadsheet import (
    build_preview_html,
    build_xlsx,
    create_spreadsheet,
    get_spreadsheet,
    update_spreadsheet,
    validate_content_json,
)
from apps.agent.tools.spreadsheet_content import revalidate_xlsx_content_json
from apps.agent.tools.table_style_tokens import COLORS
from apps.chat.models import Conversation
from apps.files.models import File
from apps.files.services import open_file_stream, serialize_file_for_ui

User = get_user_model()


def invoke_tool(tool, tool_call_id, **kwargs):
    result = tool.invoke(
        {
            "type": "tool_call",
            "name": tool.name,
            "id": tool_call_id,
            "args": kwargs,
        }
    )
    return result.content if hasattr(result, "content") else result


@pytest.fixture
def user(db):
    return User.objects.create_user(username="sheetuser", password="pass")


@pytest.fixture
def conversation(user):
    return Conversation.objects.create(user=user)


@pytest.fixture
def sample_content():
    return {
        "title": "Desglose regional",
        "sheets": [
            {
                "name": "Revenue",
                "headers": ["Region", "Revenue", "MoM %"],
                "rows": [
                    [
                        "EMEA",
                        {"value": "486,200", "align": "right"},
                        {"value": "+18.2%", "tone": "success", "align": "right"},
                    ],
                    {
                        "style": "total",
                        "cells": [
                            "Total",
                            {"value": "1,284,920", "align": "right"},
                            {"value": "+12.4%", "tone": "success", "align": "right"},
                        ],
                    },
                ],
            }
        ],
    }


@pytest.mark.django_db
class TestSpreadsheetTool:
    def test_validate_content_json(self, sample_content):
        result = validate_content_json(sample_content["title"], sample_content["sheets"])
        assert result["title"] == "Desglose regional"
        assert result["format"] == "xlsx"
        assert len(result["sheets"]) == 1
        assert result["sheets"][0]["name"] == "Revenue"

    def test_validate_requires_title(self, sample_content):
        with pytest.raises(ValueError, match="title is required"):
            validate_content_json("", sample_content["sheets"])

    def test_validate_requires_headers(self):
        with pytest.raises(ValueError, match="sheet requires headers"):
            validate_content_json("Test", [{"name": "Sheet1", "headers": [], "rows": []}])

    def test_build_xlsx_returns_bytes(self, sample_content):
        content = validate_content_json(sample_content["title"], sample_content["sheets"])
        xlsx_bytes = build_xlsx(content)
        assert isinstance(xlsx_bytes, bytes)
        assert len(xlsx_bytes) > 100
        assert xlsx_bytes[:2] == b"PK"

    def test_build_xlsx_readable_by_openpyxl(self, sample_content):
        content = validate_content_json(sample_content["title"], sample_content["sheets"])
        workbook = load_workbook(BytesIO(build_xlsx(content)))
        ws = workbook["Revenue"]
        assert ws.cell(row=1, column=1).value == "Region"
        assert ws.cell(row=2, column=1).value == "EMEA"
        assert ws.cell(row=3, column=1).value == "Total"

    def test_build_preview_html(self, sample_content):
        content = validate_content_json(sample_content["title"], sample_content["sheets"])
        html = build_preview_html(content)
        assert "ay-sheet-preview" in html
        assert "Revenue" in html
        assert "EMEA" in html
        assert "ay-sheet-preview__tab" in html

    def test_build_preview_html_grid_columns_match_header_count(self, sample_content):
        content = validate_content_json(sample_content["title"], sample_content["sheets"])
        html = build_preview_html(content)
        assert "--ay-sheet-cols: 38px repeat(3, minmax(72px, max-content))" in html

    def test_create_spreadsheet(self, user, conversation, sample_content):
        set_agent_context(conversation, user)
        result = json.loads(
            invoke_tool(
                create_spreadsheet,
                "tc-create-sheet",
                title=sample_content["title"],
                sheets=sample_content["sheets"],
                filename="regional.xlsx",
            )
        )
        assert result["ok"] is True
        assert result["action"] == "created"
        file_obj = File.objects.get(id=result["file_id"])
        assert file_obj.original_name == "regional.xlsx"
        assert file_obj.format_key == "xlsx"
        assert file_obj.content_json["format"] == "xlsx"

    def test_get_spreadsheet(self, user, conversation, sample_content):
        set_agent_context(conversation, user)
        created = json.loads(
            invoke_tool(
                create_spreadsheet,
                "tc-get-sheet",
                title=sample_content["title"],
                sheets=sample_content["sheets"],
            )
        )
        result = json.loads(invoke_tool(get_spreadsheet, "tc-read", file_id=created["file_id"]))
        assert result["ok"] is True
        assert result["content_json"]["title"] == "Desglose regional"

    def test_update_spreadsheet(self, user, conversation, sample_content):
        set_agent_context(conversation, user)
        created = json.loads(
            invoke_tool(
                create_spreadsheet,
                "tc-update-sheet-1",
                title=sample_content["title"],
                sheets=sample_content["sheets"],
            )
        )
        updated_sheets = [
            {
                "name": "Revenue",
                "headers": ["Region", "Revenue"],
                "rows": [["EMEA", "500,000"]],
            }
        ]
        result = json.loads(
            invoke_tool(
                update_spreadsheet,
                "tc-update-sheet-2",
                file_id=created["file_id"],
                sheets=updated_sheets,
            )
        )
        assert result["ok"] is True
        assert result["action"] == "updated"

    def test_update_spreadsheet_rejects_context_file(self, user, conversation):
        from apps.files.parsers import parse_upload
        from apps.files.services import save_uploaded_file
        from apps.files.tests.test_parsers_xlsx import build_sample_xlsx_bytes

        set_agent_context(conversation, user)
        parsed = parse_upload(build_sample_xlsx_bytes(), "ventas.xlsx")
        file_obj = save_uploaded_file(
            conversation=conversation,
            user=user,
            original_name="ventas.xlsx",
            file_bytes=build_sample_xlsx_bytes(),
            parsed=parsed,
        )
        result = json.loads(
            invoke_tool(
                update_spreadsheet,
                "tc-update-context",
                file_id=str(file_obj.id),
                sheets=[
                    {
                        "name": "Revenue",
                        "headers": ["Region"],
                        "rows": [["EMEA"]],
                    }
                ],
            )
        )
        assert result["ok"] is False
        assert "contexto del usuario" in result["error"]

    def test_serialize_file_for_ui(self, user, conversation, sample_content):
        set_agent_context(conversation, user)
        created = json.loads(
            invoke_tool(
                create_spreadsheet,
                "tc-serialize",
                title=sample_content["title"],
                sheets=sample_content["sheets"],
            )
        )
        file_obj = File.objects.get(id=created["file_id"])
        payload = serialize_file_for_ui(file_obj, user=user)
        assert payload["kind"] == "sheet"
        assert payload["ext"] == "XLSX"
        assert payload["role"] == "deliverable"
        assert payload["meta"] == "Spreadsheet · 1 hoja"
        assert payload["preview_url"].endswith("/preview/")

    def test_open_file_stream_rebuilds_xlsx(self, user, conversation, sample_content):
        set_agent_context(conversation, user)
        created = json.loads(
            invoke_tool(
                create_spreadsheet,
                "tc-download",
                title=sample_content["title"],
                sheets=sample_content["sheets"],
            )
        )
        file_obj = File.objects.get(id=created["file_id"])
        stream = open_file_stream(file_obj)
        with ZipFile(stream) as archive:
            assert any(name.startswith("xl/") for name in archive.namelist())

    def test_revalidate_xlsx_content_json(self, sample_content):
        content = validate_content_json(sample_content["title"], sample_content["sheets"])
        restored = revalidate_xlsx_content_json(content)
        assert restored["format"] == "xlsx"
        assert restored["sheets"][0]["headers"] == ["Region", "Revenue", "MoM %"]

    def test_revalidate_xlsx_preserves_role(self, sample_content):
        content = validate_content_json(sample_content["title"], sample_content["sheets"])
        content["role"] = "context"
        content["source"] = "upload"
        restored = revalidate_xlsx_content_json(content)
        assert restored["role"] == "context"
        assert restored["source"] == "upload"

    def test_normalize_sheet_style_defaults(self, sample_content):
        content = validate_content_json(sample_content["title"], sample_content["sheets"])
        style = content["sheets"][0]["style"]
        assert style == {"striped": True, "header_fill": "muted"}

    def test_normalize_sheet_style_custom(self):
        content = validate_content_json(
            "Test",
            [
                {
                    "name": "Data",
                    "style": {"striped": False, "header_fill": "accent_light"},
                    "headers": ["A"],
                    "rows": [["1"]],
                }
            ],
        )
        assert content["sheets"][0]["style"] == {
            "striped": False,
            "header_fill": "accent_light",
        }

    def test_validate_rejects_invalid_fill(self):
        with pytest.raises(ValueError, match="cell fill"):
            validate_content_json(
                "Test",
                [
                    {
                        "name": "Data",
                        "headers": ["A"],
                        "rows": [[{"value": "x", "fill": "purple"}]],
                    }
                ],
            )

    def test_build_xlsx_applies_cell_and_header_fills(self):
        content = validate_content_json(
            "Styled",
            [
                {
                    "name": "Styled",
                    "style": {"striped": True, "header_fill": "muted"},
                    "headers": ["Metric", "Change"],
                    "rows": [
                        [
                            "Revenue",
                            {
                                "value": "+18%",
                                "tone": "success",
                                "fill": "success_light",
                                "align": "right",
                            },
                        ],
                        {
                            "style": "total",
                            "cells": ["Total", {"value": "100", "align": "right"}],
                        },
                    ],
                }
            ],
        )
        workbook = load_workbook(BytesIO(build_xlsx(content)))
        ws = workbook["Styled"]
        header_fill = ws.cell(row=1, column=1).fill.start_color.rgb
        assert header_fill.endswith(COLORS["bg_muted"])
        success_fill = ws.cell(row=2, column=2).fill.start_color.rgb
        assert success_fill.endswith(COLORS["success_bg"])
        total_fill = ws.cell(row=3, column=1).fill.start_color.rgb
        assert total_fill.endswith(COLORS["bg_muted"])
        assert ws.cell(row=2, column=2).font.color.rgb.endswith(COLORS["success_border"])

    def test_build_preview_html_includes_fill_and_striped_classes(self):
        content = validate_content_json(
            "Styled",
            [
                {
                    "name": "Styled",
                    "style": {"striped": True, "header_fill": "muted"},
                    "headers": ["Metric", "Change"],
                    "rows": [
                        [
                            "Revenue",
                            {"value": "+18%", "fill": "success_light", "align": "right"},
                        ],
                        {
                            "style": "total",
                            "cells": ["Total", {"value": "100", "align": "right"}],
                        },
                    ],
                }
            ],
        )
        html = build_preview_html(content)
        assert "ay-sheet-preview__cell--fill-muted" in html
        assert "ay-sheet-preview__cell--fill-success_light" in html
        assert "ay-sheet-preview__cell--striped" in html
        assert "ay-sheet-preview__cell--total" in html
